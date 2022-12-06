#PROYECTO DE PYTHON
#CLAUDIA ESTEFANIA TINTAYA QUISPE
#Realizar un reporte de excel automatico
import openpyxl

wb = openpyxl.Workbook()

import openpyxl
wb = openpyxl.Workbook()
hoja = wb.active
print(f'Hoja activa: {hoja.title}')

hoja.title = "Valores"
print(f'Hoja activa: {wb.active.title}')


# Añade la hoja 'Hoja' al final (por defecto)
hoja1 = wb.create_sheet("Hoja")

# Añade la hoja 'Hoja' en la primera posición. Como el nombre
# 'Hoja' ya existe, le añade el número 1 al final del nombre
hoja2 = wb.create_sheet("Hoja", 0)

# Añade la hoja 'Otra hoja' en la posición 1
wb.create_sheet(index=1, title="Otra hoja")

# Muestra los nombres de las hojas
print(wb.sheetnames)
['Hoja1', 'Otra hoja', 'Valores', 'Hoja']


origen = wb.active
nueva = wb.copy_worksheet(origen)


hoja = wb.active  # Es la hoja que está en el índice 0
print(f'Hoja activa: {hoja.title}')
hoja = wb['Otra hoja']
wb.active = hoja
print(f'Hoja activa: {wb.active.title}')


print(wb.sheetnames)
['Hoja1', 'Otra hoja', 'Valores', 'Hoja']
for hoja in wb:
    print(hoja.title)


wb = openpyxl.Workbook()
hoja = wb.active
a1 = hoja["A1"]
print(a1.value)


b2 = hoja.cell(row=2, column=2)
print(b2.value)


# 1.- Asignando el valor directamente a la celda
hoja["A1"] = 10
a1 = hoja["A1"]
print(a1.value)


# 2.- Usando la notación fila, columna con el argumento value
b1 = hoja.cell(row=1, column=2, value=20)
print(b1.value)


# 3.- Actualizando la propiedad value de una celda
c1 = hoja.cell(row=1, column=3)
c1.value = 30
print(c1.value)


productos = [
    ('p_1', 'a859', 1500, 9.95),
    ('p_2', 'b125', 600, 4.95),
    ('p_3', 'c764', 200, 19.95),
    ('p_4', 'd399', 2000, 49.95)
]


productos = [
    ('p_1', 'a859', 1500, 9.95),
    ('p_2', 'b125', 600, 4.95),
    ('p_3', 'c764', 200, 19.95),
    ('p_4', 'd399', 2000, 49.95)
]
wb = openpyxl.Workbook()
hoja = wb.active
# Crea la fila del encabezado con los títulos
#Ref=Referencia
hoja.append(('Nombre', 'Ref', 'Stock', 'Precio'))
for producto in productos:
    # producto es una tupla con los valores de un producto 
    hoja.append(producto)
#Guardando libro excel en python
wb.save('productos.xlsx')




